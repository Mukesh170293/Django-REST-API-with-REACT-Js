import csv
from django.shortcuts import render
from rest_framework import viewsets
from rest_framework import permissions
from .models import Judge,Project,Student,JudgeAssignment
from .serializers import *
from rest_framework.decorators import api_view
from django.http.response import JsonResponse
from rest_framework.response import Response
from rest_framework.parsers import JSONParser
from rest_framework import status
from django.http import HttpResponse
from django.urls import reverse
from django.http import HttpResponseRedirect
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from openpyxl import load_workbook
from string import ascii_uppercase
from scoring import convertStrToNum as c
from django.views.decorators.csrf import csrf_exempt
from rest_framework.renderers import JSONRenderer
from rest_framework import views
from rest_framework.response import Response
from scoring.forms.scoring_form import ScoringForm
#Judgetab 
#completed
@api_view(['GET', 'POST'])
def display_judges(request):
    items = Judge.objects.all()
    judge_list = []
    for judge in items:
        judgestatus = JudgeAssignment.objects.filter(JudgeId=judge.JudgeId)
        assign_judge_list = []
        graded = 0
        for ja in judgestatus:
            assign_judge_list.append(ja.JudgeId)
            if ja.RawScore != 0:
                graded = graded + 1

        projects_assigned = len(assign_judge_list)
        if projects_assigned > 0:
            judge_list.append({"Name": str(judge.Name),
                        "JudgeId": judge.JudgeId,
                        "projects_assigned": projects_assigned,
                        "graded": int(graded),
                        "projects_assigned_half": projects_assigned / 2 }
                        )
    if request.method == 'GET':
        judge = judge_list
        Judge_Serializer = JudgedisplaySerializer(judge, many=True)
        return JsonResponse(Judge_Serializer.data, safe=False)
    elif request.method == 'POST':
        serializer = JudgedisplaySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    


# Student tab
# completed
@api_view(['GET', 'POST']) 
def display_students(request):
    student_items = Student.objects.all()
    if request.method == 'GET':
        student = Student.objects.all()
        Student_serializer = StudentSerializer(student, many=True)
        return JsonResponse(Student_serializer.data, safe=False)
    elif request.method == 'POST':
        serializer = StudentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#Display Project
#Completed
@api_view(['GET', 'POST']) 
def displayprojects(request):
    items = Project.objects.all()
    project_list = []
    for project in items:
        projectstatus = JudgeAssignment.objects.filter(ProjectId=project.ProjectId)
        assign_project_list = []
        graded = 0
        for ja in projectstatus:
            assign_project_list.append(ja.ProjectId)
            if ja.RawScore != 0:
                graded = graded + 1

        judges_assigned = len(assign_project_list)
        if judges_assigned > 0:
            project_list.append({"ProjectId": project.ProjectId,
                                 "TableId": project.TableId,
                                 "ProjectTitle": project.ProjectTitle,
                                 "ProjectCategory": project.ProjectCategory,
                                 "AvgScore": project.AvgScore,
                                 "Rank": project.Rank,
                                 "Zscore": project.Zscore,
                                 "ZscoreRank": project.ZscoreRank,
                                 "Avg_01": project.Avg_01,
                                 "Avg_01Rank": project.Avg_01Rank,
                                 "ScaledScore": project.ScaledScore,
                                 "ScaledRank": project.ScaledRank,
                                 "Scaledz": project.Scaledz,
                                 "IsefScore": project.IsefScore,
                                 "IsefRank": project.IsefRank,
                                 "CategoryRank": project.CategoryRank,
                                 "FairRank": project.FairRank,
                                 "judges_assigned": int(judges_assigned),
                                 "graded": int(graded),
                                 "judges_assigned_half": judges_assigned / 2})
    if request.method == 'GET':
        project = project_list
        Project_Serializer = ProjectdisplaySerializer(project, many=True)
        return JsonResponse(Project_Serializer.data, safe=False)
    elif request.method == 'POST':
        serializer = ProjectdisplaySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Judge assignment tab
# Completed
@api_view(['GET', 'POST'])
def display_judge_assignments(request):
    ja_items = JudgeAssignment.objects.all()

    if request.method == 'GET':
        ja_items = JudgeAssignment.objects.all()
        Judge_serializer = JudgeAssignmentSerializer(ja_items, many=True)
        return JsonResponse(Judge_serializer.data, safe=False)
    elif request.method == 'POST':
        serializer = JudgeAssignmentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Completed
# Import data tab
def import_student(sheet):
    for s in range(5, 66):
        project_id = sheet['D'+str(s)].value
        ids = [sheet['E'+str(s)].value, sheet['F'+str(s)].value,
               sheet['G'+str(s)].value]
        School = sheet['H'+str(s)].value

        for id in ids:
            if id is not None:
                s = Student(id, School, project_id)
                s.save()

def import_project(sheet):
    for s in range(5, 66):
        project_id = sheet['D'+str(s)].value
        table_id = sheet['C'+str(s)].value
        description = sheet['L'+str(s)].value
        project_title = sheet['J'+str(s)].value
        project_category = sheet['K'+str(s)].value
        avg_score = None
        rank = None
        z_score = None
        z_score_rank = None
        avg_01 = None
        avg_01_rank = None
        scaled_score = None
        scaled_rank = None
        scaled_z = None
        isef_score = None
        isef_rank = None

        p = Project(project_id, c.int1(table_id), description, project_title,
                    project_category, c.float1(avg_score), c.int1(rank),
                    c.float1(z_score), c.float1(scaled_score),
                    c.float1(scaled_rank), c.float1(scaled_z),
                    c.float1(isef_score), c.int1(isef_rank))
        p.save()

def import_judge(sheet):
    alphabet = ascii_uppercase
    judge_id = []
    name = []
    for C in alphabet[13:26]:
        j_id = sheet[str(C)+str(4)].value
        j_n = sheet[str(C)+str(3)].value
        if '-' in j_id and j_n:
            continue
        if j_id not in judge_id:
            judge_id.append(j_id)
        if j_n not in name:
            name.append(j_n)
    for C1 in alphabet[0:8]:
        for C2 in alphabet:
            j_id = sheet[str(C1)+str(C2)+str(4)].value
            j_n = sheet[str(C1)+str(C2)+str(3)].value
            if j_id is None:
                break
            if '-' in j_id and j_n:
                continue
            if j_id not in judge_id:
                judge_id.append(j_id)
            if j_n not in name:
                name.append(j_n)

    for j_id, j_n in zip(judge_id, name):
        j = Judge(j_id, j_n)
        j.save()


def import_judge_assignment(sheet):
    goal_score = 0
    plan_score = 0
    action_score = 0
    result_analysis_score = 0
    communication_score = 0
    raw_score = 0
    id = 0
    for row in sheet.iter_rows(min_row=5, min_col=13, max_row=56, max_col=220):
        for cell in row:
            if cell.value:
                if cell.value.strip():
                    id = id + 1
                    j_id = sheet.cell(row = 4, column = cell.column).value
                    if '-' in j_id:
                        continue
                    p_id = sheet.cell(row = cell.row, column = 4).value
                    ja = JudgeAssignment(id, j_id, p_id,
                                          goal_score,
                                          plan_score,
                                          action_score,
                                          result_analysis_score,
                                          communication_score,
                                          raw_score)
                    ja.save()



def import_data(request):
    wb = load_workbook(filename = request, data_only = True)
    sheet = wb.active
    import_judge(sheet)
    import_project(sheet)
    import_student(sheet)
    import_judge_assignment(sheet)


def import_file(request):
    if request.method=='POST':
        try:
            file = request.FILES['document']
            import_data(file)
            messages.success(request, "Data imported successfully")
        except Exception:
            messages.success(request, "Please choose file first")
    return render(request, "import.html")

@api_view(['POST'])
def import_file_upload(request):
    import_data(request.data['file'])
    return Response(None, status=status.HTTP_201_CREATED)

# Export Judge assignment
# Complete
def export_judge_assignment(request):
    response = HttpResponse(content_type='text/csv')
    new_file = "judge_assignments.csv"
    response['Content-Disposition'] = 'attachment; filename=judge.csv'

    writer = csv.writer(response)
    writer.writerow(['Judge ID', 'Project ID', 'goal score', 'plan_score',
                     'action score', 'result Analyis score',
                     'communication score', 'Raw Score', 'z_score', 'rank'])

    jas = JudgeAssignment.objects.all().values_list('JudgeId', 'ProjectId',
                                                     'GoalScore',
                                                     'PlanScore',
                                                     'ActionScore',
                                                     'ResultAnalysisScore',
                                                     'CommunicationScore',
                                                     'RawScore',
                                                     'Zscore', 'Rank')
    for ja in jas:
        writer.writerow(list(ja))

    return response

# Export project
# Completed
def export_project(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="projects.csv"'

    writer = csv.writer(response)
    writer.writerow(['project_id', 'table_id', 'project_title',
                     'project_category', 'avg_score', 'rank', 'z_score',
                     'z_score_rank', 'avg_01', 'avg_01_rank', 'scaled_score',
                     'scaled_rank', 'scaled_z', 'isef_score', 'isef_rank',
                     'category_rank', 'fair_rank'])

    projs = Project.objects.all().values_list('ProjectId', 'TableId',
                                              'ProjectTitle',
                                              'ProjectCategory',
                                              'AvgScore', 'Rank',
                                              'Zscore', 'ZscoreRank',
                                              'Avg_01', 'Avg_01Rank',
                                              'ScaledScore', 'ScaledRank',
                                              'Scaledz', 'IsefScore',
                                              'IsefRank', 'CategoryRank',
                                              'FairRank')
    for proj in projs:
        writer.writerow(list(proj))

    return response

# Remove all data tab
# Completed
@api_view(['DELETE'])
def remove_all_data(request):
    if request.method == "DELETE":  
        Judge.objects.all().delete()
        Project.objects.all().delete()
        Student.objects.all().delete()
        JudgeAssignment.objects.all().delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


# Top projects tab
# Completed
@api_view(['GET', 'POST'])
def display_top_projects(request):
    top_all_list = []
    top_all_categories = []
    top_all = list(Project.objects.order_by('FairRank').
                   exclude(FairRank__isnull=True)[:5])
    for tls in top_all:
        stu = Student.objects.filter(ProjectId=tls.ProjectId)
        fns = " "
        for st in stu:
            fns = fns + " " + st.Id
        top_all_list.append({"ProjectId": tls.ProjectId,
                             "ProjectTitle": tls.ProjectTitle,
                             "ProjectCategory": tls.ProjectCategory,
                             "CategoryRank": tls.CategoryRank,
                             "FairRank": tls.FairRank,
                             "student_names": fns})

    projects = list(Project.objects.all())
    categories = []
    for project in projects:
        if not project.ProjectCategory in categories:
            categories.append(project.ProjectCategory)

    top_all_studcat = []
    for category in categories:
        top_in_category = Project.objects.filter(ProjectCategory=category).order_by('CategoryRank')[:5]
        top_in_categorys = []
        for top in top_in_category:
            students = list(Student.objects.filter(ProjectId=top.ProjectId))
            sn = ""
            for s in students:
                sn = sn + " " + s.Id
            top_in_categorys.append({"ProjectId": top.ProjectId,
                                     "ProjectTitle": top.ProjectTitle,
                                     "ProjectCategory": top.ProjectCategory,
                                     "CategoryRank": top.CategoryRank,
                                     "FairRank": top.FairRank,
                                     "student_names": sn})
        top_all_categories.append(top_in_categorys)

    if request.method == 'GET':
        top = top_in_categorys
        displaytopprojectsSerializer = display_top_projects_Serializer(top, many=True)
        return JsonResponse(displaytopprojectsSerializer.data, safe=False)
    elif request.method == 'POST':
        serializer = display_top_projects_Serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        




# display unassigned tab
# Display Complete
@api_view(['GET', 'POST'])
def removeassignedprojectassignments(request):
    items = Project.objects.all()
    item = Judge.objects.all()
    project_list = []
    assigned_projects = []
    for project in items:
        projectstatus = JudgeAssignment.objects.filter(ProjectId=project.ProjectId)
        assign_project_list = []
        graded = 0

        for ja in projectstatus:
            assign_project_list.append(ja.ProjectId)
            if ja.RawScore != 0:
                graded = graded + 1

        judges_assigned = len(assign_project_list)
        if judges_assigned > 0:
            assigned_projects.append({"ProjectId": project.ProjectId,
                                      "ProjectTitle": project.ProjectTitle})
    
    if request.method == 'GET':
        assigned_projects_list = assigned_projects
        project_Serializer = UnassignProjectSerializer(assigned_projects_list, many=True)
        return JsonResponse(project_Serializer.data, safe=False)
    elif request.method == 'POST':
        serializer = UnassignProjectSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Display Complete
@api_view(['GET', 'POST'])
def removeassignedjudgeassignments(request):
    items_judge = Judge.objects.all()
    judge_list = []
    unassigned_judges = []
    assigned_projects = []
    assigned_judges = []

    for judge in items_judge:
        judgestatus = JudgeAssignment.objects.filter(JudgeId=judge.JudgeId)
        assign_judge_list = []
        graded = 0

        for ja in judgestatus:
            assign_judge_list.append(ja.JudgeId)
            if ja.RawScore != 0:
                graded = graded + 1
        projects_assigned = len(assign_judge_list)
        if projects_assigned > 0:
            assigned_judges.append({"Name": str(judge.Name),
                                    "JudgeId": judge.JudgeId})

    if request.method == 'GET':
        assigned_judges_list = assigned_judges
        Judge_Serializer = JudgeSerializer(assigned_judges_list, many=True)
        return JsonResponse(Judge_Serializer.data, safe=False)
    elif request.method == 'POST':
        serializer = JudgeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Assignment
def display_unasigned(request):
    context = {}
    judges_dict = unassigned_judges()
    projects_dict = unassigned_projects()
    context['button'] = 'button'
    context['unassigned_judges'] = judges_dict['unassigned_judges']
    context['unassigned_projects'] = projects_dict['unassigned_projects']

    return render(request, 'assigned.html', context)


#remove judge tab
# complete
@api_view(['GET','DELETE'])
def deletejudgeassignment(request,pk):
    try:
        judge = Judge.objects.get(pk=pk)
    except Judge.DoesNotExist:
        return HttpResponse(status=404)

    if request.method == 'GET':
        serializer = JudgeSerializer(judge)
        return Response(serializer.data)

    elif request.method == 'DELETE':
        judge.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# Remove Project
# complete
@api_view(['GET','DELETE'])
def deleteprojectassignment(request,pk):
    try:
        project = Project.objects.get(pk=pk)
    except Project.DoesNotExist:
        return HttpResponse(status=404)

    if request.method == 'GET':
        serializer = ProjectSerializer(project)
        return Response(serializer.data)

    elif request.method == 'DELETE':
        project.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# scoring
def add_score(request):
    if request.method == 'POST':
        form = ScoringForm(request.POST)
        if form.is_valid():
            form.save()
    else:
        form = ScoringForm()
    ja_items = JudgeAssignment.objects.all()
    context = {
        'form': form,
        'ja_items': ja_items,
    }
    return render(request, 'scoring.html', context)

